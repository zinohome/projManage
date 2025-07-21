#!/usr/bin/python3
# -*- coding:utf-8 -*-
"""
@author: ibmzhangjun@139.com
@file: excel_to_text.py
@time: 2025/7/21 下午2:34
@desc: 
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import argparse
import os

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import argparse
import os

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import argparse
import os


def is_integer(value):
    """判断一个值是否为整数（包括整数类型和可以无损转换为整数的浮点数）"""
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    return False


def truncate_string(s, max_length=240):
    """将字符串截断为指定长度，超长时添加省略号标记"""
    if not isinstance(s, str) or len(s) <= max_length:
        return s
    return s[:max_length - 3] + '...'  # 保留前237个字符，添加...


def convert_excel_to_text(input_file, output_file=None, sheet_name=None):
    """
    将 Excel 文件中的所有单元格转换为文本格式

    参数:
    input_file (str): 输入 Excel 文件路径
    output_file (str, optional): 输出 Excel 文件路径。默认为 None，将在原文件名后添加 '_text'
    sheet_name (str, optional): 要处理的表名。默认为 None，表示处理所有表
    """
    # 若未指定输出文件，在原文件名后添加 '_text'
    if output_file is None:
        base_name, ext = os.path.splitext(input_file)
        output_file = f"{base_name}_text{ext}"

    # 读取 Excel 文件
    try:
        xls = pd.ExcelFile(input_file)

        # 获取所有表名
        sheet_names = xls.sheet_names
        if sheet_name is not None:
            if sheet_name not in sheet_names:
                raise ValueError(f"指定的表名 '{sheet_name}' 不存在")
            sheet_names = [sheet_name]

        # 创建新的工作簿
        wb = openpyxl.Workbook()
        # 删除默认创建的工作表
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 处理每个工作表
        for sheet in sheet_names:
            # 读取数据
            df = xls.parse(sheet)

            # 创建新的工作表
            new_sheet = wb.create_sheet(title=sheet)

            # 添加表头
            header = df.columns.tolist()
            for col_idx, value in enumerate(header, 1):
                cell = new_sheet.cell(row=1, column=col_idx)
                # 截断表头字符串
                cell.value = truncate_string(str(value)) if pd.notna(value) else ""
                cell.number_format = "@"

            # 添加数据行
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = new_sheet.cell(row=row_idx, column=col_idx)
                    # 处理不同类型的值
                    if pd.isna(value):
                        cell.value = ""
                    elif isinstance(value, str):
                        # 截断字符串值
                        cell.value = truncate_string(value)
                    elif is_integer(value):
                        # 处理整数
                        cell.value = str(int(value))
                    elif isinstance(value, float):
                        # 处理浮点数
                        cell.value = str(value)
                    elif isinstance(value, pd.Timestamp):
                        # 保留原始日期时间格式
                        cell.value = value.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(value, bool):
                        cell.value = str(value)
                    else:
                        # 截断其他类型转换后的字符串
                        cell.value = truncate_string(str(value)) if value is not None else ""

                    # 设置单元格格式为文本
                    cell.number_format = "@"

        # 保存工作簿
        wb.save(output_file)
        print(f"已成功将文件保存为: {output_file}")

    except Exception as e:
        print(f"处理文件时出错: {str(e)}")


def main():
    parser = argparse.ArgumentParser(description='将 Excel 文件中的所有单元格转换为文本格式')
    parser.add_argument('input_file', help='输入 Excel 文件路径')
    parser.add_argument('-o', '--output', help='输出 Excel 文件路径')
    parser.add_argument('-s', '--sheet', help='要处理的表名，默认为所有表')
    parser.add_argument('--max-length', type=int, default=240,
                        help='字符串最大长度，超过时自动截断并添加省略号 (默认: 240)')

    args = parser.parse_args()

    # 使用全局变量传递最大长度参数
    global MAX_STRING_LENGTH
    MAX_STRING_LENGTH = args.max_length

    convert_excel_to_text(args.input_file, args.output, args.sheet)


if __name__ == "__main__":
    main()